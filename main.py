import pandas as pd
import openpyxl
import decode
import visualize

global arquivo_excel

def add_excel(lista, nome_coluna):
    workbook = openpyxl.load_workbook(arquivo_excel)
    sheet = workbook.active

    nova_coluna_valores = lista

    # Definir o cabeçalho da nova coluna
    cabecalho_nova_coluna = nome_coluna
    sheet.cell(row=1, column=sheet.max_column + 1, value=cabecalho_nova_coluna)

    # Adicionar os valores da lista à nova coluna
    for i, valor in enumerate(nova_coluna_valores, start=2):
        sheet.cell(row=i, column=sheet.max_column, value=valor)

    # Salvar as alterações no arquivo
    workbook.save(arquivo_excel)

    print("Coluna adicionada com sucesso!")



# Substitua 'nome_do_arquivo' pelo caminho para o arquivo que você deseja ler
arquivo = 'config/21.bin'
dic = decode.decodificar_arquivo(arquivo)
bytes = decode.file_to_byte_array(arquivo)


"""
Classe usada para manipular o dado recebido em algo que seja legível e entendível

-> Fazer com que a classe receba o valor direto do arquivo e a partir de uma lambda function 
processe os dados 

-> A taxa de amostragem dos dados nem sempre é a mesme, o mesmo vale pro tempo de medida, por 
isso temos que fazer as funções de visualização levando isso em consideração, uma possível 
solução pra isso é fazer um scatter plot quando formos plotar os dados pelo tempo (Ex:. (apps_1, apps2) x tempo)
e podemos usar uma lookup table quando formos plotar um dado por outro
"""
class dado:
    def __init__(self, nome, origem):
        self.nome = nome
        self.origem = origem
        self.dados = []
        self.tempos = []
        self.unidade = ''

    def __getitem__(self, index):
        return self.dados[index] 
    
    def scatter(self):
        return [(self.tempos[i], self.dados[i]) for i in range(len(self.dados))]
    



#-------------------------------------------------------
#-------------------------------------------------------

apps1_bruto = dado("apps1_bruto", dic)
apps1_bruto.dados = [(i["dados"][1] << 8) | (i["dados"][0]) for i in dic if i["id"] == 3]
apps1_bruto.tempos = [(i["tick"]) for i in dic if i["id"] == 3]

apps1_tensao = dado("apps1_tensão", dic)
apps1_tensao.dados = [float((i["dados"][1] << 8) | (i["dados"][0])) * (3.3/4095.0) for i in dic if i["id"] == 3]
apps1_tensao.tempos = [(i["tick"]) for i in dic if i["id"] == 3]

apps2_bruto = dado("apps2_bruto", dic)
apps2_bruto.dados = [(i["dados"][3] << 8) | (i["dados"][2]) for i in dic if i["id"] == 3]
apps2_bruto.tempos = [(i["tick"]) for i in dic if i["id"] == 3]

apps2_tensao = dado("apps2_tensão", dic)
apps2_tensao.dados = [float((i["dados"][3] << 8) | (i["dados"][2])) * (3.3/4095.0) for i in dic if i["id"] == 3]
apps2_tensao.tempos = [(i["tick"]) for i in dic if i["id"] == 3]

bse1_bruto = dado("bse1_bruto", dic)
bse1_bruto.dados = [(i["dados"][5] << 8) | (i["dados"][4]) for i in dic if i["id"] == 3]
bse1_bruto.tempos = [(i["tick"]) for i in dic if i["id"] == 3]

bse1_tensao = dado("bse1_tensão", dic)
bse1_tensao.dados = [float((i["dados"][5] << 8) | (i["dados"][4])) * (3.3/4095.0) for i in dic if i["id"] == 3]
bse1_tensao.tempos = [(i["tick"]) for i in dic if i["id"] == 3]

volante_bruto = dado("volante_bruto", dic)
volante_bruto.dados = [(i["dados"][7] << 8) | (i["dados"][6])  for i in dic if i["id"] == 3]
volante_bruto.tempos = [(i["tick"]) for i in dic if i["id"] == 3]


#-------------------------------------------------------
#-------------------------------------------------------

pack_current = dado("PACK_current", dic)
pack_current.dados = [ ((i["dados"][0] << 8) | (i["dados"][1])) / 10 for i in dic if i["id"] == 59]
pack_current.tempos = [(i["tick"]) for i in dic if i["id"] == 59]

## Filtro passa baixa

import numpy as np
from scipy.signal import butter,filtfilt

def butter_lowpass_filter(data, cutoff, fs, order):
    nyq = 0.5 * fs  # Nyquist Frequency
    normal_cutoff = cutoff / nyq
    # Get the filter coefficients 
    b, a = butter(order, normal_cutoff, btype='low', analog=False)
    y = filtfilt(b, a, data)
    return y

pack_current.dados = butter_lowpass_filter(pack_current.dados, 2, 1000 / (pack_current.tempos[1] - pack_current.tempos[0]), 2)

## Filtro média movel

tam_grupo = 10 # deslocamento (uma semana útil)
medias_moveis=[]


for i in range(len(pack_current.dados)):
    if pack_current.dados[i] > 100:
        pack_current.dados[i] = 100


i = 0
# Calcular as médias móveis e armazená-las em uma lista:
while i < len(pack_current.dados) - tam_grupo + 1:
    grupo = pack_current.dados[i : i + tam_grupo]
    media_grupo = sum(grupo) / tam_grupo
    medias_moveis.append(media_grupo)
    i +=1

pack_current.dados = medias_moveis



for i in range(tam_grupo - 1):
    pack_current.dados.append(0)

######################
#-------------------------------------------------------
#-------------------------------------------------------

apps_erro = dado("APPS_erro", dic)
bse_erro = dado("BSE_erro", dic)
bppc_erro = dado("BPPC_erro", dic)

for i in dic:
    if i["id"] == 288:
        apps_erro.tempos.append(i["tick"])
        bse_erro.tempos.append(i["tick"])
        bppc_erro.tempos.append(i["tick"])
        if i["dados"][0] == 0:
            apps_erro.dados.append(0)
            bse_erro.dados.append(0)
            bppc_erro.dados.append(0)
        elif i["dados"][0] == 1:
            apps_erro.dados.append(1)
            bse_erro.dados.append(0)
            bppc_erro.dados.append(0)
        elif i["dados"][0] == 4:
            apps_erro.dados.append(0)
            bse_erro.dados.append(0)
            bppc_erro.dados.append(1)
        elif i["dados"][0] == 5:
            apps_erro.dados.append(1)
            bse_erro.dados.append(0)
            bppc_erro.dados.append(1)



# apps_erro.dados = [i["dados"][0] for i in dic if i["id"] == 288]
# apps_erro.tempos = [(i["tick"]) for i in dic if i["id"] == 288]

# bse_erro.dados = [i["dados"][0] for i in dic if i["id"] == 288]
# bse_erro.tempos = [(i["tick"]) for i in dic if i["id"] == 288]

# bppc_erro.dados = [i["dados"][0] for i in dic if i["id"] == 288]
# bppc_erro.tempos = [(i["tick"]) for i in dic if i["id"] == 288]


#-------------------------------------------------------
#-------------------------------------------------------

bse2_bruto = dado("bse2_bruto", dic)
bse2_bruto.dados = [(i["dados"][4] << 8) | (i["dados"][5]) for i in dic if i["id"] == 5 and i["dlc"] == 6]
bse2_bruto.tempos = [(i["tick"]) for i in dic if i["id"] == 5 and i["dlc"] == 6]


#-------------------------------------------------------
#-------------------------------------------------------


pack_soc = dado("pack_soc", dic)
pack_soc.dados = [ (i["dados"][1]) for i in dic if i["id"] == 1714]
pack_soc.tempos = [(i["tick"]) for i in dic if i["id"] == 1714]


#-------------------------------------------------------
#-------------------------------------------------------




arquivo_excel = 'Corrida_'

for letra in arquivo:
    if letra.isdigit():
        arquivo_excel = arquivo_excel + letra

arquivo_excel  = arquivo_excel + '.xlsx'



tabela_id3 = pd.DataFrame({
    'ticks_id3': apps1_tensao.tempos,
    'apps1_bruto': apps1_bruto.dados,
    'apps2_bruto': apps2_bruto.dados,
    'bse1_bruto': bse1_bruto.dados,
    'volante_bruto': volante_bruto.dados
})


#visualize.plotar_tabela(dic)
#visualize.plotar_grafico_composto([ apps1_tensao, apps2_tensao ])
visualize.plotar_grafico_composto([ pack_current ])
visualize.mostrar()


# tabela_id3.to_excel(arquivo_excel, index=False)
# add_excel(bse2_bruto.tempos, 'tick_id5')
# add_excel(bse2_bruto.dados, 'bse2_bruto')
# add_excel(pack_soc.tempos, 'tick_id1714')
# add_excel(pack_soc.dados, 'pack_soc')
# add_excel(apps_erro.tempos, 'tick_id288')
# add_excel(apps_erro.dados, 'apps_erro')
# add_excel(bse_erro.dados, 'bse_erro')
# add_excel(bppc_erro.dados, 'bppc_erro')




